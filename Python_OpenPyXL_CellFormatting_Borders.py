
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Style
from openpyxl import Workbook



#import openpyxl
#from openpyxl import Workbook
#from openpyxl.styles import Border, Side, Color
#from openpyxl.styles import Style
import os
from DefaultItems import DataFolder


thin_border = Border(left=Side(border_style='thin'),
                         right=Side(border_style='thin'),
                         top=Side(border_style='thin'),
                         bottom=Side(border_style='thin'))

my_style = Style(border=thin_border)


filename = 'Open Work_20200220A.xlsx'

pnf_in  = os.path.join(r"c:\users\ken.stephani\desktop", 'Open Work_20200220A.xlsx')
pnf_out = os.path.join(r"c:\users\ken.stephani\desktop", 'Open Work_20200220A_out.xlsx')

print(pnf_in)
print(pnf_out)

wb = openpyxl.load_workbook(pnf_in)
print(wb)
sn = wb.get_sheet_names()
print(sn)
ws = wb.get_active_sheet()

#    sheet = wb['Page 1']

#    print(sheet['B2'].value)

ws.cell(row = 2, column = 2).style=my_style
wb.save(pnf_in)
    #print(pnf_out)
'''

    sheet = wb.get_sheet_by_name('Metadata - Countries')
    title = sheet.title
    print(title)
    print(sheet.max_row)
    print(sheet.max_column)
    #sheet.title = 'New Title'
    wb.save(filename)
    #wb.get_sheet_names()
'''

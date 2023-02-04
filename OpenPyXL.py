import openpyxl
import os
from DefaultItems import DataFolder

filename = 'gd1.xlsx'

pnf = os.path.join(DataFolder, filename)
print("8")

wb = openpyxl.load_workbook(pnf)
sn = wb.get_sheet_names()
print(sn)
sheet = wb.get_sheet_by_name('Metadata - Countries')
title = sheet.title
print(title)
print(sheet.max_row)
print(sheet.max_column)
#sheet.title = 'New Title'
wb.save(filename)
#wb.get_sheet_names()

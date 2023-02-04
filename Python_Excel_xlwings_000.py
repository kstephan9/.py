import openpyxl
import pprint
import os
from openpyxl.utils  import   get_column_letter, column_index_from_string


from DefaultItems import DataFolder, OutputFolder, Raw_2Folder

filename = 'SurveyTest_000.xlsx'

pnf = os.path.join(Raw_2Folder, filename)

#wb = openpyxl.load_workbook(pnf)
wb = openpyxl.open_workbook(pnf)
sn = wb.sheetnames
print(wb.sheetnames)

sheet1 = wb.sheet_by_index(0)

print("sheet1['A1'].value: ",sheet1['A1'].value)

for i, value_tuple in enumerate(sheet1.rows):
    print(i, value_tuple[1].value)



